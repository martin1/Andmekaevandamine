#coding=utf-8

'''
Created on Sep 14, 2012

@author: martin
'''

from openpyxl.reader.excel import load_workbook
from string import lower

file_name = "/home/martin/dev/files/tshekid.xlsx"

wb = load_workbook(file_name)
sheet = wb.worksheets[0]

i = 1 #xlsx row iterator
data = [] #list for data rows
row = "" #data row (string)

while sheet.cell(column=1, row=i).value:
    buyer = sheet.cell(column=0, row=i).value
    
    product = sheet.cell(column=1, row=i).value.replace(", ", "_")
    product = product.replace(" ", "_")
    product = lower(product)
    
    next_buyer = sheet.cell(column=0, row=i+1).value
    
    if buyer != next_buyer: #end of one buyer's items
        row = row + " " + product
        data.append(row)
        row=""
    else: #previous buyer's items continue
        if row =="":
            row += product
        else:
            row = row + " " + product
    i+=1
    
with open("/home/martin/dev/files/input2.txt", "w") as apriori:
    for item in data:
        apriori.write(item + "\n")





